#!/usr/bin/env python3
"""Coleta dados de mercado para o Painel Soja.

Roda dentro do GitHub Actions (runner com internet livre). Cada coletor é
independente: se um falhar, a seção correspondente mantém os dados da
execução anterior (lidos de data/data.json) e o erro fica registrado em
`errors` — o painel nunca quebra por causa de uma fonte fora do ar.

Fontes:
  - Yahoo Finance (não oficial): futuros CBOT (inclusive curva de
    vencimentos), Brent, Ibovespa e câmbio de fallback
  - AwesomeAPI: dólar e euro comercial (limita IPs do Actions; há fallback)
  - Banco Central (Olinda/PTAX e SGS): PTAX, Selic, CDI, IPCA
  - Notícias Agrícolas: indicadores Cepea/Esalq, prêmio de porto e balcão
  - USDA/NASS Crop Progress (via API da biblioteca Cornell): safra dos EUA
  - CONAB (série histórica de grãos): safra do Brasil
  - CME Group (FTP público ftp.cmegroup.com): volume de calls/puts de soja
  - USDA/FAS PSD (api.fas.usda.gov, chave em USDA_FAS_KEY): oferta e demanda
  - USDA/FAS ESR (mesma chave): embarques e vendas semanais dos EUA
  - Open-Meteo: previsão de chuva em cidades produtoras do Centro-Oeste
  - RSS: Google News, Canal Rural, G1 Agronegócios, Notícias Agrícolas
"""

import datetime as dt
import email.utils
import html
import json
import os
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

import requests

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "data" / "data.json"
TIMEOUT = 25
TAG_RE = re.compile(r"<[^>]+>")
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
}


def now_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")


# ------------------------------------------------------- cadência e frescor
#
# `updated_at` diz quando o robô rodou, e é a mesma hora para todas as seções.
# Não diz nada sobre o frescor do dado: um embarque referente a 13/08 e uma
# cotação de minutos atrás carregavam o mesmo carimbo. Quem responde "de
# quando é este número" é `observado_em`, a data da própria observação.

CADENCIA = {
    "quotes": "intradiaria",
    "curve": "intradiaria",
    "fx": "intradiaria",
    "options": "diaria",
    "cepea": "diaria",
    "basis": "diaria",
    "weather": "diaria",
    "news": "continua",
    "rates": "mista",  # PTAX/Selic/CDI diários, IPCA mensal
    "safra": "mista",  # EUA semanal (segundas), Brasil mensal
    "embarques": "semanal",  # divulgado nas quintas, ~1 semana de defasagem
    "sd": "mensal",  # WASDE
}


def data_br_iso(s):
    """`25/08/2026` -> `2026-08-25`. None se não reconhecer."""
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", (s or "").strip())
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else None


def observado_em(nome: str, sec):
    """Data da observação de uma seção, quando a fonte a informa.

    Só devolve o que dá para afirmar. Seção intradiária não tem data de
    observação separada — o dado é de agora, e aí `updated_at` já basta.
    """
    if not isinstance(sec, dict):
        return None
    if nome == "cepea":
        return data_br_iso((sec.get("indicador") or {}).get("data"))
    if nome == "basis":
        return data_br_iso(sec.get("indicador_data"))
    if nome == "safra":
        return (sec.get("us") or {}).get("week_ending")
    if nome == "embarques":
        datas = [i.get("semana_ref") for i in sec.get("itens") or [] if i.get("semana_ref")]
        return max(datas) if datas else None
    if nome == "sd":
        # O PSD informa o mês do levantamento, não o dia. Precisão de mês.
        return sec.get("vintage_iso")
    return None


def anota_cadencia(sections: dict) -> None:
    """Carimba cada seção com cadência, data da observação e idade em dias."""
    hoje = dt.date.today()
    for nome, sec in sections.items():
        if not isinstance(sec, dict):
            continue
        sec["cadencia"] = CADENCIA.get(nome, "desconhecida")
        obs = observado_em(nome, sec)
        if not obs:
            continue
        sec["observado_em"] = obs
        try:
            partes = [int(p) for p in obs.split("-")]
            d = dt.date(partes[0], partes[1], partes[2] if len(partes) > 2 else 1)
            sec["idade_dias"] = (hoje - d).days
            sec["precisao_observacao"] = "dia" if len(partes) > 2 else "mes"
        except (ValueError, IndexError):
            pass


def get(url: str, **kwargs) -> requests.Response:
    r = requests.get(url, headers=HEADERS, timeout=TIMEOUT, **kwargs)
    r.raise_for_status()
    return r


# ---------------------------------------------------------------- cotações

YAHOO_SYMBOLS = [
    ("ZS=F", "Soja CBOT", "¢/bushel"),
    ("ZC=F", "Milho CBOT", "¢/bushel"),
    ("ZM=F", "Farelo de soja", "US$/t curta"),
    ("ZL=F", "Óleo de soja", "¢/lb"),
    ("ZW=F", "Trigo CBOT", "¢/bushel"),
    ("BZ=F", "Petróleo Brent", "US$/barril"),
    ("^BVSP", "Ibovespa", "pontos"),
]


def fetch_yahoo_symbol(symbol: str, range_: str = "1mo") -> dict:
    last_err = None
    for host in ("query1", "query2"):
        try:
            r = get(
                f"https://{host}.finance.yahoo.com/v8/finance/chart/{symbol}",
                params={"range": range_, "interval": "1d"},
            )
            result = r.json()["chart"]["result"][0]
            meta = result["meta"]
            quote = result["indicators"]["quote"][0]
            closes = [round(c, 4) for c in quote.get("close") or [] if c is not None]
            price = meta.get("regularMarketPrice")
            if price is None and closes:
                price = closes[-1]
            prev = closes[-2] if len(closes) >= 2 else meta.get("chartPreviousClose")
            change_pct = (
                round((price / prev - 1) * 100, 2) if price and prev else None
            )
            volume = meta.get("regularMarketVolume")
            if volume is None:
                vols = [v for v in quote.get("volume") or [] if v]
                volume = vols[-1] if vols else None
            return {
                "price": price,
                "prev_close": prev,
                "change_pct": change_pct,
                "closes": closes,
                "volume": volume,
            }
        except Exception as e:  # noqa: BLE001 - tenta o próximo host
            last_err = e
    raise RuntimeError(f"{symbol}: {last_err}")


_YAHOO_AUTH = None


def yahoo_auth():
    """Sessão autenticada do Yahoo (cookie + crumb), uma por execução.

    O Yahoo limita por IP e o Actions bate no 429 com facilidade, então vale
    pagar o par cookie+crumb só uma vez e reaproveitar entre os coletores.
    """
    global _YAHOO_AUTH
    if _YAHOO_AUTH is None:
        s = requests.Session()
        s.headers.update(HEADERS)
        s.get("https://fc.yahoo.com", timeout=TIMEOUT)
        crumb = s.get(
            "https://query1.finance.yahoo.com/v1/test/getcrumb", timeout=TIMEOUT
        ).text.strip()
        # Sob limite de taxa o Yahoo devolve 200 com o texto "Too Many
        # Requests" no lugar do crumb — sem espaço nem "<" não é crumb válido.
        if not crumb or "<" in crumb or " " in crumb:
            raise RuntimeError(f"crumb do Yahoo inválido: {crumb[:40]!r}")
        _YAHOO_AUTH = (s, crumb)
    return _YAHOO_AUTH


def fetch_yahoo_quotes(symbols: list) -> dict:
    """Cotações em lote pelo endpoint /v7/finance/quote.

    É o único caminho gratuito que devolve `openInterest` (posições em
    aberto) por contrato e `underlyingSymbol` (qual contrato está por trás de
    um símbolo contínuo como `ZC=F`). Exige cookie + crumb.
    """
    s, crumb = yahoo_auth()
    r = s.get(
        "https://query1.finance.yahoo.com/v7/finance/quote",
        params={"symbols": ",".join(symbols), "crumb": crumb},
        timeout=TIMEOUT,
    )
    r.raise_for_status()
    return {q["symbol"]: q for q in r.json()["quoteResponse"]["result"]}


# Séries guardadas por símbolo. O painel fatia essas três para montar os
# períodos de 1 dia a 5 anos sem ir buscar nada de novo no clique.
YAHOO_SERIES = (
    ("intraday", "5d", "60m"),
    ("daily", "1y", "1d"),
    ("weekly", "5y", "1wk"),
)


def fetch_yahoo_series(symbol: str, range_: str, interval: str):
    """Devolve ({t: [...], c: [...]}, meta) para um símbolo e granularidade."""
    last_err = None
    for host in ("query1", "query2"):
        try:
            r = get(
                f"https://{host}.finance.yahoo.com/v8/finance/chart/{symbol}",
                params={"range": range_, "interval": interval},
            )
            res = r.json()["chart"]["result"][0]
            ts = res.get("timestamp") or []
            closes = res["indicators"]["quote"][0].get("close") or []
            pares = [(t, round(c, 2)) for t, c in zip(ts, closes) if c is not None]
            return {"t": [p[0] for p in pares], "c": [p[1] for p in pares]}, res["meta"]
        except Exception as e:  # noqa: BLE001
            last_err = e
    raise RuntimeError(f"{symbol} {range_}/{interval}: {last_err}")


def close_ate(serie: dict, alvo: int):
    """Último fechamento em ou antes de `alvo` (timestamp unix)."""
    anterior = None
    for t, c in zip(serie["t"], serie["c"]):
        if t <= alvo:
            anterior = c
        else:
            break
    return anterior


HORIZONTES = (
    ("semana", 7),
    ("mes", 30),
    ("trimestre", 91),
    ("ano", 365),
    ("cinco_anos", 1826),
)


def fetch_yahoo_completo(symbol: str) -> dict:
    series, meta = {}, None
    erros = []
    for nome, rng, itv in YAHOO_SERIES:
        try:
            s, m = fetch_yahoo_series(symbol, rng, itv)
            if s["c"]:
                series[nome] = s
                meta = meta or m
        except Exception as e:  # noqa: BLE001
            erros.append(str(e))
    if "daily" not in series:
        raise RuntimeError("; ".join(erros) or f"{symbol}: sem série diária")

    diaria = series["daily"]
    semanal = series.get("weekly", diaria)
    price = (meta or {}).get("regularMarketPrice")
    if price is None:
        price = diaria["c"][-1]
    prev = diaria["c"][-2] if len(diaria["c"]) >= 2 else None

    agora = diaria["t"][-1]
    variacoes = {}
    if prev:
        variacoes["dia"] = round((price / prev - 1) * 100, 2)
    for nome, dias in HORIZONTES:
        alvo = agora - dias * 86400
        # Usa a série diária quando ela alcança o horizonte; senão a semanal.
        # Só reporta se alguma delas de fato cobre o período — melhor omitir
        # do que comparar contra o começo da série e inventar a variação.
        for base in (diaria, semanal):
            if base["t"] and base["t"][0] <= alvo:
                ref = close_ate(base, alvo)
                if ref:
                    variacoes[nome] = round((price / ref - 1) * 100, 2)
                break

    volume = (meta or {}).get("regularMarketVolume")
    return {
        "price": price,
        "prev_close": prev,
        "change_pct": variacoes.get("dia"),
        "closes": diaria["c"][-22:],  # faísca do cartão
        "volume": volume,
        "variacoes": variacoes,
        "series": series,
    }


# Códigos de mês dos futuros (padrão das bolsas americanas).
MESES_CBOT = {
    "F": "jan", "G": "fev", "H": "mar", "J": "abr",
    "K": "mai", "M": "jun", "N": "jul", "Q": "ago",
    "U": "set", "V": "out", "X": "nov", "Z": "dez",
}


def rotulo_contrato(symbol: str):
    """`ZCZ26.CBT` -> `dez/26`. None se o símbolo não tiver esse formato."""
    m = re.match(r"^[A-Z]{2,3}([FGHJKMNQUVXZ])(\d{2})$", (symbol or "").split(".")[0])
    return f"{MESES_CBOT[m.group(1)]}/{m.group(2)}" if m else None


def aplica_quote_v7(item: dict, q: dict) -> dict:
    """Nomeia o contrato do contínuo e corrige a variação do dia.

    Um símbolo contínuo (`ZC=F`) não é um contrato: aponta para o vencimento
    ativo e troca sozinho. A série do /v8/chart é emendada sem ajuste, então
    no dia da rolagem o preço salta do contrato velho para o novo e a
    variação calculada pela série vira o spread entre vencimentos — foi assim
    que o milho apareceu com +6,46% em 25/08/2026, quando o ZC=F passou de
    setembro para dezembro.

    O /v7/quote compara o contrato com o fechamento dele mesmo, imune à
    emenda, e ainda diz em `underlyingSymbol` qual contrato está cotando.
    """
    if not q:
        return item
    contrato = q.get("underlyingSymbol")
    if contrato:
        item["contrato"] = contrato
        item["contrato_rotulo"] = rotulo_contrato(contrato)
    change = q.get("regularMarketChangePercent")
    prev = q.get("regularMarketPreviousClose")
    if change is not None:
        item["change_pct"] = round(change, 2)
        item["variacoes"] = {**item.get("variacoes", {}), "dia": round(change, 2)}
    if prev is not None:
        item["prev_close"] = prev
    return item


def collect_quotes() -> dict:
    items, failed = [], []

    quotes = {}
    try:
        quotes = fetch_yahoo_quotes([s for s, _, _ in YAHOO_SYMBOLS])
    except Exception as e:  # noqa: BLE001 - sem crumb, fica só a série emendada
        failed.append(f"lote /v7/quote: {e}")

    for symbol, name, unit in YAHOO_SYMBOLS:
        try:
            q = fetch_yahoo_completo(symbol)
            item = {"symbol": symbol, "name": name, "unit": unit, **q}
            items.append(aplica_quote_v7(item, quotes.get(symbol)))
        except Exception as e:  # noqa: BLE001
            failed.append(f"{symbol}: {e}")
    if not items:
        raise RuntimeError("; ".join(failed))
    return {"updated_at": now_iso(), "items": items, "failed": failed}


# ---------------------------------------------------------------- câmbio

def collect_fx() -> dict:
    fx = {"updated_at": now_iso()}
    try:
        data = get(
            "https://economia.awesomeapi.com.br/json/last/USD-BRL,EUR-BRL"
        ).json()

        def pick(key):
            d = data[key]
            return {
                "bid": float(d["bid"]),
                "ask": float(d["ask"]),
                "high": float(d["high"]),
                "low": float(d["low"]),
                "pct_change": float(d["pctChange"]),
            }

        fx["usdbrl"] = pick("USDBRL")
        fx["eurbrl"] = pick("EURBRL")
        fx["source"] = "AwesomeAPI"
    except Exception as e:  # noqa: BLE001 - a AwesomeAPI limita IPs do Actions (429)
        # Fallback: Yahoo Finance. Atenção — `BRL=X` não é o dólar comercial:
        # é o spot negociado 24 h no exterior. Depois que o mercado brasileiro
        # fecha os dois descolam, e a variação do dia chega a sair com sinal
        # oposto (em 25/08/2026 o comercial fechou -0,29% e o Yahoo marcava
        # +0,19%). Por isso o painel precisa saber de onde veio o número.
        fx["awesomeapi_error"] = str(e)
        for key, symbol in (("usdbrl", "BRL=X"), ("eurbrl", "EURBRL=X")):
            q = fetch_yahoo_symbol(symbol)
            fx[key] = {
                "bid": q["price"],
                "ask": None,
                "high": None,
                "low": None,
                "pct_change": q["change_pct"],
            }
        fx["source"] = "Yahoo Finance"
        fx["spot_24h"] = True

    try:
        end = dt.date.today()
        start = end - dt.timedelta(days=10)
        fmt = "%m-%d-%Y"
        url = (
            "https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/"
            "CotacaoDolarPeriodo(dataInicial=@dataInicial,dataFinalCotacao=@dataFinalCotacao)"
            f"?@dataInicial='{start.strftime(fmt)}'&@dataFinalCotacao='{end.strftime(fmt)}'"
            "&$top=1&$orderby=dataHoraCotacao%20desc&$format=json"
        )
        v = get(url).json()["value"][0]
        fx["ptax"] = {
            "compra": v["cotacaoCompra"],
            "venda": v["cotacaoVenda"],
            "data": v["dataHoraCotacao"][:10],
        }
    except Exception as e:  # noqa: BLE001 - PTAX é complementar
        fx["ptax"] = None
        fx["ptax_error"] = str(e)
    return fx


# ---------------------------------------------------------------- juros/macro

def sgs(codigo: int, ultimos: int = 1) -> list:
    r = get(
        f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{codigo}"
        f"/dados/ultimos/{ultimos}?formato=json"
    )
    return r.json()


def collect_rates() -> dict:
    out = {"updated_at": now_iso()}
    selic = sgs(432)[-1]
    out["selic"] = float(selic["valor"])
    try:
        cdi = sgs(4389)[-1]  # CDI anualizado base 252, % a.a.
        out["cdi"] = float(cdi["valor"])
    except Exception:  # noqa: BLE001
        out["cdi"] = None
    ipca = sgs(433, 12)
    out["ipca_mes"] = {"valor": float(ipca[-1]["valor"]), "data": ipca[-1]["data"]}
    acc = 1.0
    for p in ipca:
        acc *= 1 + float(p["valor"]) / 100
    out["ipca_12m"] = round((acc - 1) * 100, 2)
    return out


# ---------------------------------------------------------------- físico

NA_SOJA = "https://www.noticiasagricolas.com.br/cotacoes/soja"

# Praças do Centro-Oeste e portos, na ordem de exibição. O site traz o nome
# da fonte entre parênteses depois da praça, então casamos por prefixo.
PRACAS = [
    "Sorriso/MT",
    "Rondonópolis/MT",
    "Primavera do Leste/MT",
    "Rio Verde/GO",
    "Jataí/GO",
    "Campo Grande/MS",
    "Maracaju/MS",
    "São Gabriel do Oeste/MS",
    "Oeste da Bahia/BA",
    "Porto Paranaguá (disponível)",
    "Porto Santos/SP",
]


def br_float(s: str):
    s = (s or "").strip().replace("+", "")
    if not s or "cota" in s.lower():
        return None
    try:
        return float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def parse_na_tabelas(page: str) -> dict:
    """Extrai as tabelas de cotação da página, indexadas pelo título."""
    blocos = re.findall(
        r'<h2>\s*<a[^>]*title="([^"]+)".*?<table[^>]*>(.*?)</table>', page, re.S
    )
    tabelas = {}
    for titulo, tabela in blocos:
        linhas = []
        for ln in re.findall(r"<tr>(.*?)</tr>", tabela, re.S):
            celulas = [
                " ".join(html.unescape(TAG_RE.sub("", c)).split())
                for c in re.findall(r"<td[^>]*>(.*?)</td>", ln, re.S)
            ]
            if len(celulas) >= 2:
                linhas.append(celulas)
        tabelas[html.unescape(titulo).strip()] = linhas
    return tabelas


def collect_cepea() -> dict:
    """Indicadores CEPEA e mercado físico via Notícias Agrícolas.

    O site do CEPEA passou a exigir verificação da Cloudflare e devolve 403
    para qualquer acesso automatizado. O Notícias Agrícolas republica os
    mesmos indicadores citando a fonte, em HTML simples, e ainda traz o
    prêmio de porto e os preços de balcão nas praças produtoras.
    """
    tabelas = parse_na_tabelas(get(NA_SOJA).text)

    def indicador(titulo: str, nome: str):
        for linha in tabelas.get(titulo, []):
            if len(linha) >= 3 and re.match(r"\d{2}/\d{2}/\d{4}", linha[0]):
                valor = br_float(linha[1])
                if valor is not None:
                    return {
                        "nome": nome,
                        "data": linha[0],
                        "valor": valor,
                        "var_dia_pct": br_float(linha[2]),
                    }
        return None

    indicadores = [
        i
        for i in (
            indicador(
                "Indicador da Soja ESALQ/B3 - Paranaguá",
                "Indicador Soja ESALQ/B3 — Paranaguá",
            ),
            indicador(
                "Indicador da Soja Cepea/Esalq - Paraná",
                "Indicador Soja CEPEA/ESALQ — Paraná",
            ),
        )
        if i
    ]
    if not indicadores:
        raise RuntimeError("indicadores da soja não encontrados na página")

    out = {
        "updated_at": now_iso(),
        "fonte": "Cepea/Esalq, via Notícias Agrícolas",
        "indicador": indicadores[0],
        "indicadores": indicadores,
    }

    # A tabela é cotada em US$/bushel (cabeçalho "US$/Bu"), não em centavos.
    premio = [
        {
            "mes": ln[0],
            "usd_bu": br_float(ln[1]),
            "cents_bu": round(br_float(ln[1]) * 100, 1),
            "var": br_float(ln[2]) if len(ln) > 2 else None,
        }
        for ln in tabelas.get("Prêmio Soja Paranaguá/PR", [])
        if "/" in ln[0] and br_float(ln[1]) is not None
    ]
    if premio:
        out["premio_paranagua"] = premio

    fisico = []
    for ln in tabelas.get("Soja - Mercado Físico", []):
        if len(ln) < 3:
            continue
        praca = next((p for p in PRACAS if ln[0].startswith(p)), None)
        valor = br_float(ln[1])
        if praca and valor is not None:
            # A tabela nomeia quem cotou — "Rio Verde/GO (Comigo)",
            # "Rondonópolis/MT (Samir Rosa Assessoria Comercial)". São preços
            # de balcão de cooperativas, sindicatos e corretoras, e é por isso
            # que divergem de consultorias como a Safras & Mercado: base,
            # prazo de pagamento e ponto de entrega são outros. Guardar o
            # agente é o que torna a diferença legível em vez de suspeita.
            # Só o que vem depois do nome da praça: em "Porto Paranaguá
            # (disponível)" o parêntese é parte do próprio nome, não o agente.
            resto = ln[0].strip()[len(praca) :]
            m = re.search(r"\(([^)]+)\)\s*$", resto)
            agente = m.group(1).strip() if m else None
            fisico.append(
                {
                    "praca": praca,
                    "valor": valor,
                    "var_dia_pct": br_float(ln[2]),
                    "agente": agente,
                }
            )
    if fisico:
        fisico.sort(key=lambda f: PRACAS.index(f["praca"]))
        out["fisico"] = fisico
    return out


# ---------------------------------------------------------------- clima

CITIES = [
    ("Sorriso · MT", -12.55, -55.72),
    ("Sinop · MT", -11.86, -55.50),
    ("Rio Verde · GO", -17.79, -50.93),
    ("Dourados · MS", -22.22, -54.81),
]


def collect_weather() -> dict:
    """Previsão de 7 dias e chuva acumulada nos 30 dias anteriores.

    O acumulado recente é o que define se a janela de plantio abre: sem
    umidade no solo o produtor não semeia, mesmo dentro do calendário.
    """
    lats = ",".join(str(c[1]) for c in CITIES)
    lons = ",".join(str(c[2]) for c in CITIES)
    r = get(
        "https://api.open-meteo.com/v1/forecast",
        params={
            "latitude": lats,
            "longitude": lons,
            "daily": "precipitation_sum,temperature_2m_max,temperature_2m_min",
            "forecast_days": 7,
            "past_days": 31,
            "timezone": "America/Sao_Paulo",
        },
    )
    payload = r.json()
    if isinstance(payload, dict):
        payload = [payload]
    hoje = dt.date.today().isoformat()
    cities = []
    for (name, _, _), res in zip(CITIES, payload):
        d = res["daily"]
        datas = d["time"]
        corte = next((i for i, t in enumerate(datas) if t >= hoje), len(datas))
        passado = [p for p in d["precipitation_sum"][:corte] if p is not None]
        cities.append(
            {
                "name": name,
                "dates": datas[corte:],
                "precip": d["precipitation_sum"][corte:],
                "tmax": d["temperature_2m_max"][corte:],
                "tmin": d["temperature_2m_min"][corte:],
                "precip_30d": round(sum(passado[-30:]), 1),
            }
        )
    return {"updated_at": now_iso(), "cities": cities}


# ---------------------------------------------------------------- safra

SOY_ACTIVITIES = {
    "planted": "Plantio",
    "emerged": "Emergência",
    "blooming": "Floração",
    "setting pods": "Formação de vagens",
    "dropping leaves": "Queda de folhas",
    "harvested": "Colheita",
}


def collect_safra_us() -> dict:
    """Crop Progress semanal do USDA/NASS via API da biblioteca Cornell.

    O zip da publicação traz `prog_all_tables.csv` com todas as tabelas;
    filtramos as de soja e extraímos a linha-resumo "18 States".
    """
    import csv
    import io
    import zipfile

    meta = get(
        "https://usda.library.cornell.edu/api/v1/release/findByIdentifier/CropProg",
        params={"latest": "true"},
    ).json()
    release = meta["results"][0]
    zip_url = next(f for f in release["files"] if f.endswith(".zip"))
    z = zipfile.ZipFile(io.BytesIO(get(zip_url).content))
    text = z.read("prog_all_tables.csv").decode("cp1252", "replace")

    tables = {}
    for row in csv.reader(io.StringIO(text)):
        if len(row) >= 2:
            tables.setdefault(row[0], []).append(row[1:])

    def num(s):
        s = (s or "").strip()
        if s in ("", "-"):
            return 0
        try:
            return int(s)
        except ValueError:
            return None

    out = {
        "released": release["release_datetime"][:10],
        "week_ending": None,
        "coverage": None,
        "progress": [],
        "condition": None,
    }
    for rows in tables.values():
        titles = [r[1] for r in rows if r[0] == "t" and len(r) > 1]
        title = next(
            (t for t in titles if "Soybean" in t and not t.startswith("Crop Progress")),
            None,
        )
        if not title:
            continue
        m = re.search(r"Week Ending ([A-Z][a-z]+ \d+, \d{4})", title)
        if m:
            try:
                out["week_ending"] = dt.datetime.strptime(
                    m.group(1), "%B %d, %Y"
                ).date().isoformat()
            except ValueError:
                out["week_ending"] = m.group(1)
        note = next((t for t in titles if "States planted" in t), None)
        if note and not out["coverage"]:
            c = re.search(r"These (\d+) States planted (\d+)%", note)
            if c:
                out["coverage"] = {
                    "states": int(c.group(1)),
                    "pct_area": int(c.group(2)),
                }
        data = {
            r[1]: [num(c) for c in r[2:]]
            for r in rows
            if r[0] == "d" and len(r) > 2 and r[1]
        }
        if "Condition" in title:
            cur = data.get("18 States")
            if cur and len(cur) >= 5:
                cond = {
                    "dist": dict(zip(("vp", "p", "f", "g", "e"), cur)),
                    "gex": cur[3] + cur[4],
                }
                for label, key in (
                    ("Previous week", "gex_prev_week"),
                    ("Previous year", "gex_prev_year"),
                ):
                    v = data.get(label)
                    if v and len(v) >= 5:
                        cond[key] = v[3] + v[4]
                out["condition"] = cond
        else:
            # Colunas: mesma semana do ano passado, semana anterior,
            # semana atual, média de 5 anos
            cur = data.get("18 States")
            if cur and len(cur) >= 4:
                activity = title.replace("Soybeans", "").split("Selected")[0]
                activity = re.sub(r"[^A-Za-z ]", " ", activity).strip()
                label = SOY_ACTIVITIES.get(activity.lower(), activity)
                out["progress"].append(
                    {
                        "label": label,
                        "prev_year": cur[0],
                        "prev_week": cur[1],
                        "current": cur[2],
                        "avg_5y": cur[3],
                    }
                )
    if not out["progress"] and not out["condition"]:
        raise RuntimeError("nenhuma tabela de soja encontrada no Crop Progress")
    ordem = list(SOY_ACTIVITIES.values())
    out["progress"].sort(
        key=lambda p: ordem.index(p["label"]) if p["label"] in ordem else 99
    )
    return out


def fase_safra_br(hoje: dt.date = None) -> dict:
    """Em que ponto do calendário da soja brasileira estamos.

    Não existe fonte gratuita e estruturada com o percentual plantado por
    semana (a CONAB publica em painel Power BI e a AgRural, em release);
    o calendário abaixo, somado à chuva acumulada, dá a leitura do momento.
    """
    d = hoje or dt.date.today()
    m, dia = d.month, d.day
    if (m == 9 and dia >= 16) or m == 10:
        return {
            "fase": "Plantio",
            "detalhe": "janela de semeadura aberta no Centro-Oeste",
        }
    if m in (11, 12):
        return {
            "fase": "Desenvolvimento",
            "detalhe": "lavoura em desenvolvimento; plantio tardio no Sul",
        }
    if m in (1, 2, 3):
        return {"fase": "Colheita", "detalhe": "colheita da safra de verão"}
    if m in (4, 5):
        return {
            "fase": "Fim de colheita",
            "detalhe": "encerramento da colheita e plantio da safrinha",
        }
    return {
        "fase": "Entressafra",
        "detalhe": "comercialização; vazio sanitário até meados de setembro",
    }


def collect_safra_br() -> dict:
    """Série histórica de grãos da CONAB (área, produção e produtividade)."""
    r = get(
        "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/"
        "SerieHistoricaGraos.txt"
    )
    por_safra = {}
    for line in r.text.splitlines()[1:]:
        p = [c.strip() for c in line.split(";")]
        if len(p) < 8 or p[3] != "SOJA":
            continue
        agg = por_safra.setdefault(p[0], {"area": 0.0, "prod": 0.0})
        try:
            agg["area"] += float(p[5])
            agg["prod"] += float(p[6])
        except ValueError:
            continue
    if not por_safra:
        raise RuntimeError("nenhuma linha de SOJA na série da CONAB")
    safras = sorted(por_safra)
    atual, anterior = safras[-1], (safras[-2] if len(safras) > 1 else None)

    def fmt(safra):
        a = por_safra[safra]
        return {
            "safra": safra,
            "area_mil_ha": round(a["area"], 1),
            "producao_mil_t": round(a["prod"], 1),
            "produtividade_kg_ha": round(a["prod"] / a["area"] * 1000, 1)
            if a["area"]
            else None,
        }

    out = {
        "atual": fmt(atual),
        "anterior": fmt(anterior) if anterior else None,
        "ciclo": fase_safra_br(),
    }
    if anterior:
        prev = por_safra[anterior]
        cur = por_safra[atual]
        if prev["prod"]:
            out["var_producao_pct"] = round((cur["prod"] / prev["prod"] - 1) * 100, 1)
        if prev["area"]:
            out["var_area_pct"] = round((cur["area"] / prev["area"] - 1) * 100, 1)
    return out


def collect_safra() -> dict:
    out = {"updated_at": now_iso(), "failed": []}
    for key, fn in (("us", collect_safra_us), ("br", collect_safra_br)):
        try:
            out[key] = fn()
        except Exception as e:  # noqa: BLE001
            out[key] = None
            out["failed"].append(f"{key}: {e}")
    if out["us"] is None and out["br"] is None:
        raise RuntimeError("; ".join(out["failed"]))
    return out


# ---------------------------------------------------------------- curva CBOT

# Meses de vencimento na CBOT (código, mês, rótulo pt-BR). Soja e milho têm
# calendários diferentes: a soja segue o ciclo do esmagamento (jan, mar, mai,
# jul, ago, set, nov) e o milho o da colheita americana (mar, mai, jul, set,
# dez), com o dezembro como contrato da safra nova.
SOY_MONTHS = [
    ("F", 1, "jan"),
    ("H", 3, "mar"),
    ("K", 5, "mai"),
    ("N", 7, "jul"),
    ("Q", 8, "ago"),
    ("U", 9, "set"),
    ("X", 11, "nov"),
]

CORN_MONTHS = [
    ("H", 3, "mar"),
    ("K", 5, "mai"),
    ("N", 7, "jul"),
    ("U", 9, "set"),
    ("Z", 12, "dez"),
]

# O bushel é uma medida de volume: cada grão tem um peso legal próprio. Soja
# pesa 60 lb/bushel (27,2155 kg) e milho 56 lb (25,4012 kg). Usar a constante
# da soja para converter milho em R$/saca erra ~7% para menos.
CURVAS = [
    {"id": "soja", "nome": "Soja CBOT", "prefixo": "ZS", "meses": SOY_MONTHS,
     "kg_bushel": 27.2155},
    {"id": "milho", "nome": "Milho CBOT", "prefixo": "ZC", "meses": CORN_MONTHS,
     "kg_bushel": 25.4012},
]


def contratos_cbot(prefixo: str, meses: list, n: int = 7) -> list:
    """Próximos n vencimentos de um grão na CBOT a partir do mês corrente."""
    today = dt.date.today()
    # o contrato vence por volta do dia 15; depois disso pula para o próximo
    cutoff = (today.year, today.month + (1 if today.day > 15 else 0))
    out = []
    for year in range(today.year, today.year + 3):
        for code, month, label in meses:
            if (year, month) < cutoff:
                continue
            out.append(
                {
                    "symbol": f"{prefixo}{code}{year % 100:02d}.CBT",
                    "label": f"{label}/{year % 100:02d}",
                }
            )
            if len(out) >= n:
                return out
    return out


def monta_contratos(wanted: list, quotes: dict, failed: list) -> list:
    """Preenche preço/volume/OI de cada vencimento, com fallback pelo chart."""
    contracts = []
    for c in wanted:
        q = quotes.get(c["symbol"])
        if q:
            change = q.get("regularMarketChangePercent")
            contracts.append(
                {
                    "symbol": c["symbol"],
                    "label": c["label"],
                    "price": q.get("regularMarketPrice"),
                    "change_pct": round(change, 2) if change is not None else None,
                    "volume": q.get("regularMarketVolume"),
                    "open_interest": q.get("openInterest"),
                    "expires_at": (q.get("expireIsoDate") or "")[:10] or None,
                }
            )
            continue
        try:
            f = fetch_yahoo_symbol(c["symbol"], range_="5d")
            contracts.append(
                {
                    "symbol": c["symbol"],
                    "label": c["label"],
                    "price": f["price"],
                    "change_pct": f["change_pct"],
                    "volume": f.get("volume"),
                    "open_interest": None,
                    "expires_at": None,
                }
            )
        except Exception as e:  # noqa: BLE001
            failed.append(f"{c['symbol']}: {e}")

    # Spread para o vencimento anterior: positivo = carrego (o mercado paga
    # para estocar), negativo = inversão (o mercado quer o grão agora). É esse
    # degrau que aparece como "alta" quando o contínuo do Yahoo rola.
    anterior = None
    for c in contracts:
        if c["price"] is not None and anterior is not None:
            c["spread_anterior"] = round(c["price"] - anterior, 2)
        else:
            c["spread_anterior"] = None
        if c["price"] is not None:
            anterior = c["price"]
    return contracts


def collect_curve() -> dict:
    failed = []
    pedidos = {c["id"]: contratos_cbot(c["prefixo"], c["meses"]) for c in CURVAS}

    # Um crumb e uma requisição só para as duas curvas — o /v7/quote do Yahoo
    # limita por IP e o Actions bate no 429 com facilidade.
    simbolos = [w["symbol"] for lista in pedidos.values() for w in lista]
    quotes = {}
    try:
        quotes = fetch_yahoo_quotes(simbolos)
    except Exception as e:  # noqa: BLE001 - sem crumb, cai para o chart (sem OI)
        failed.append(f"lote /v7/quote: {e}")

    curvas = []
    for meta in CURVAS:
        contracts = monta_contratos(pedidos[meta["id"]], quotes, failed)
        if not contracts:
            continue
        ois = [c["open_interest"] for c in contracts if c.get("open_interest")]
        curvas.append(
            {
                "id": meta["id"],
                "nome": meta["nome"],
                "unit": "¢/bushel",
                "kg_bushel": meta["kg_bushel"],
                "contracts": contracts,
                "total_open_interest": sum(ois) if ois else None,
            }
        )

    if not curvas:
        raise RuntimeError("; ".join(failed))
    return {"updated_at": now_iso(), "curvas": curvas, "failed": failed}


def curva_de(sections: dict, id_: str) -> dict:
    """Uma curva pelo id, para quem depende de um grão específico."""
    for c in (sections.get("curve") or {}).get("curvas", []):
        if c.get("id") == id_:
            return c
    return {}


# ---------------------------------------------------------------- opções CBOT

def collect_options() -> dict:
    """Volume diário de calls e puts de soja na CBOT.

    Fonte: relatório público `daily_volume.xlsx` no FTP da CME
    (ftp.cmegroup.com), atualizado a cada pregão. O site/API da CME bloqueia
    os IPs do GitHub Actions, mas o FTP é aberto. O arquivo traz volume por
    produto (calls e puts separados); posições em aberto só existem
    agregadas por grupo, então não são publicadas por produto.
    """
    import io
    from ftplib import FTP

    from openpyxl import load_workbook

    buf, last_err = None, None
    for _ in range(3):
        try:
            ftp = FTP("ftp.cmegroup.com", timeout=60)
            ftp.login()
            buf = io.BytesIO()
            ftp.retrbinary("RETR daily_volume/daily_volume.xlsx", buf.write)
            ftp.close()
            break
        except Exception as e:  # noqa: BLE001 - o FTP da CME derruba conexões às vezes
            last_err = e
            buf = None
    if buf is None:
        raise RuntimeError(f"FTP da CME: {last_err}")

    wb = load_workbook(buf, read_only=True)
    ws = next(w for w in wb.worksheets if "by Product" in w.title)
    trade_date = None
    vol_idx = None
    wanted = {"SOYBEAN FUTURE": "futures", "SOYBEAN CALL": "calls", "SOYBEAN PUT": "puts"}
    out = {}
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c) for c in row]
        if cells and cells[0].startswith("Trade Date"):
            m = re.search(r"(\d{2})/(\d{2})/(\d{4})", cells[0])
            if m:
                trade_date = f"{m.group(3)}-{m.group(1)}-{m.group(2)}"
        if "Product Description" in cells:
            vol_idx = next(
                i for i, c in enumerate(cells) if c.replace("\n", " ").strip() == "Total Volume"
            )
            desc_idx = cells.index("Product Description")
            continue
        if vol_idx is None or len(cells) <= vol_idx:
            continue
        key = wanted.get(cells[desc_idx].strip())
        if key:
            try:
                out[key] = {"volume": int(float(cells[vol_idx]))}
            except ValueError:
                continue
    if "calls" not in out or "puts" not in out:
        raise RuntimeError("linhas SOYBEAN CALL/PUT não encontradas no daily_volume.xlsx")
    cbot = out
    if cbot["calls"]["volume"]:
        cbot["put_call_volume"] = round(
            cbot["puts"]["volume"] / cbot["calls"]["volume"], 2
        )
    return {"updated_at": now_iso(), "trade_date": trade_date, "cbot": cbot}


# ---------------------------------------------------------------- notícias

FEEDS = [
    ("Google News", "https://news.google.com/rss/search?q=soja+OR+%22mercado+de+gr%C3%A3os%22&hl=pt-BR&gl=BR&ceid=BR:pt-419"),
    ("Canal Rural", "https://www.canalrural.com.br/feed/"),
    ("G1 Agronegócios", "https://g1.globo.com/rss/g1/economia/agronegocios/"),
    ("Notícias Agrícolas", "https://www.noticiasagricolas.com.br/rss/noticias"),
]

def parse_feed(source: str, url: str) -> list:
    r = get(url)
    root = ET.fromstring(r.content)
    items = []
    for item in root.iter("item"):
        title = html.unescape(TAG_RE.sub("", item.findtext("title") or "")).strip()
        link = (item.findtext("link") or "").strip()
        if not title or not link:
            continue
        src = item.findtext("source")
        pub = item.findtext("pubDate")
        published = None
        if pub:
            try:
                published = email.utils.parsedate_to_datetime(pub).isoformat()
            except Exception:  # noqa: BLE001
                pass
        items.append(
            {
                "title": title,
                "link": link,
                "source": (src or source).strip(),
                "published_at": published,
            }
        )
        if len(items) >= 12:
            break
    return items


def collect_news() -> dict:
    items, failed, seen = [], [], set()
    for source, url in FEEDS:
        try:
            for it in parse_feed(source, url):
                key = re.sub(r"\W+", "", it["title"].lower())[:80]
                if key in seen:
                    continue
                seen.add(key)
                items.append(it)
        except Exception as e:  # noqa: BLE001
            failed.append(f"{source}: {e}")
    if not items:
        raise RuntimeError("; ".join(failed))
    items.sort(key=lambda i: i["published_at"] or "", reverse=True)
    return {"updated_at": now_iso(), "items": items[:30], "failed": failed}


# ---------------------------------------------------------------- oferta e demanda

FAS_BASE = "https://api.fas.usda.gov/api"


def fas(path: str, servico: str = "psd", timeout: int = TIMEOUT):
    """API do USDA/FAS. A chave vive no secret USDA_FAS_KEY.

    `servico` escolhe entre "psd" (balanço de oferta e demanda) e "esr"
    (Export Sales — embarques e vendas semanais).
    """
    chave = os.environ.get("USDA_FAS_KEY", "").strip()
    if not chave:
        raise RuntimeError("USDA_FAS_KEY não configurada")
    r = requests.get(
        f"{FAS_BASE}/{servico}{path}",
        headers={**HEADERS, "X-Api-Key": chave, "Accept": "application/json"},
        timeout=timeout,
    )
    r.raise_for_status()
    return r.json()


def collect_sd() -> dict:
    """Balanço de oferta e demanda da soja em grão — mundo, Brasil e EUA.

    O código da commodity e os nomes dos atributos são descobertos na hora:
    o PSD tem uma linha separada para farelo e óleo, e travar IDs numéricos
    é a forma mais fácil de acabar publicando farelo como se fosse grão.
    """
    commodities = fas("/commodities")
    grao = next(
        (
            c
            for c in commodities
            if "soybean" in c["commodityName"].lower()
            and "oilseed" in c["commodityName"].lower()
        ),
        None,
    )
    if not grao:
        raise RuntimeError("commodity 'Oilseed, Soybean' não encontrada no PSD")
    codigo = grao["commodityCode"]
    nomes = {
        a["attributeId"]: a["attributeName"].strip()
        for a in fas("/commodityAttributes")
    }

    ano = dt.date.today().year
    # O ano-safra corrente vira em setembro; antes disso o PSD ainda
    # publica o anterior como o mais recente com dados.
    anos = [ano, ano - 1]

    def balanco(path_fmt: str):
        for a in anos:
            registros = fas(path_fmt.format(ano=a))
            if not registros:
                continue
            vals, mes, ano_cal = {}, None, None
            for r in registros:
                nome = nomes.get(r["attributeId"])
                if nome and r.get("unitId") == 8:  # 1000 toneladas
                    vals[nome] = round(r["value"] / 1000, 2)  # -> Mt
                elif nome:
                    vals[nome] = r["value"]
                mes = r.get("month") or mes
                # ano civil da divulgação — não confundir com o ano-safra
                ano_cal = r.get("calendarYear") or ano_cal
            if vals:
                return a, mes, vals, ano_cal
        return None, None, {}, None

    def resumo(vals: dict, mundo: bool = False) -> dict:
        def pega(*chaves):
            for k in chaves:
                if k in vals:
                    return vals[k]
            return None

        prod = pega("Production")
        fim = pega("Ending Stocks")
        cons = pega("Domestic Consumption", "Total Dom. Cons.", "TOTAL Dom. Cons.")
        exp = pega("Exports", "Total Exports")
        # No agregado mundial a exportação é transferência entre países, não
        # uso: somá-la contaria o mesmo grão duas vezes e achataria o S/U.
        uso = (cons or 0) if mundo else (cons or 0) + (exp or 0)
        return {
            "producao": prod,
            "estoque_inicial": pega("Beginning Stocks"),
            "estoque_final": fim,
            "consumo": cons,
            "exportacao": exp,
            "importacao": pega("Imports", "Total Imports"),
            "area": pega("Area Harvested"),
            "su_pct": round(fim / uso * 100, 1) if fim is not None and uso else None,
        }

    escopos = {}
    alvos = {
        "mundo": f"/commodity/{codigo}/world/year/{{ano}}",
        "br": f"/commodity/{codigo}/country/BR/year/{{ano}}",
        "us": f"/commodity/{codigo}/country/US/year/{{ano}}",
    }
    mes_ref, ano_ref, ano_cal_ref = None, None, None
    for nome, path in alvos.items():
        a, mes, vals, ano_cal = balanco(path)
        if not vals:
            continue
        eh_mundo = nome == "mundo"
        atual = resumo(vals, eh_mundo)
        ano_ref, mes_ref = a, mes or mes_ref
        ano_cal_ref = ano_cal or ano_cal_ref

        _, _, vals_ant, _ = balanco(path.replace("{ano}", str(a - 1)))
        anterior = resumo(vals_ant, eh_mundo) if vals_ant else None
        if anterior and atual["producao"] and anterior["producao"]:
            atual["var_producao_pct"] = round(
                (atual["producao"] / anterior["producao"] - 1) * 100, 1
            )
        if anterior and atual["estoque_final"] and anterior["estoque_final"]:
            atual["var_estoque_pct"] = round(
                (atual["estoque_final"] / anterior["estoque_final"] - 1) * 100, 1
            )
        atual["anterior"] = anterior
        escopos[nome] = atual

    if not escopos:
        raise RuntimeError("PSD não devolveu dados para nenhum escopo")

    return {
        "updated_at": now_iso(),
        "commodity": grao["commodityName"].strip(),
        "safra": f"{ano_ref}/{str((ano_ref or 0) + 1)[-2:]}" if ano_ref else None,
        "vintage": mes_ref,
        # Mês do levantamento, com o ano civil da divulgação — o PSD não
        # informa o dia, então a precisão é de mês.
        "vintage_iso": (
            f"{ano_cal_ref}-{str(mes_ref).zfill(2)}" if ano_cal_ref and mes_ref else None
        ),
        "unidade": "Mt",
        "escopos": escopos,
    }


# ---------------------------------------------------------------- embarques

# Commodities do ESR e como achar a correspondente no PSD, que usa outra
# nomenclatura. O casamento é por nome, nunca por código fixo.
ESR_ALVOS = [
    ("soja", "Soybeans", lambda n: "oilseed" in n and "soybean" in n),
    ("farelo", "Soybean cake & meal", lambda n: "meal" in n and "soybean" in n),
    ("oleo", "Soybean Oil", lambda n: n.startswith("oil,") and "soybean" in n),
    ("milho", "Corn", lambda n: n == "corn"),
]
CHINA = 5700
SEMANAS_SERIE = 16


def psd_export_us(psd_commodities: list, casa, nomes_attr: dict, ano: int):
    """Projeção de exportação dos EUA no PSD, em toneladas."""
    alvo = next(
        (c for c in psd_commodities if casa(c["commodityName"].strip().lower())), None
    )
    if not alvo:
        return None, None
    registros = fas(f"/commodity/{alvo['commodityCode']}/country/US/year/{ano}")
    for r in registros:
        nome = nomes_attr.get(r["attributeId"], "")
        if nome in ("Exports", "Total Exports") and r.get("unitId") == 8:
            return r["value"] * 1000, alvo["commodityName"].strip()
    return None, alvo["commodityName"].strip()


def collect_embarques() -> dict:
    """Embarques e vendas semanais dos EUA (USDA/FAS Export Sales).

    O número da semana sozinho não diz nada — o que informa é o confronto
    com o ritmo necessário para cumprir a projeção do WASDE, com a mesma
    semana do ano anterior e com a média das últimas quatro.

    Cuidado com o ano-safra: no ESR, marketYear N vai de setembro de N-1 a
    agosto de N (a safra N-1/N); no PSD, marketYear N é a safra N/N+1. Ou
    seja, PSD = ESR - 1. Confundir os dois compara safras diferentes e o
    erro passa despercebido porque as ordens de grandeza são parecidas.
    """
    calendario = {}
    for r in fas("/datareleasedates", "esr"):
        cod = r["commodityCode"]
        if cod not in calendario or r["marketYear"] > calendario[cod]["marketYear"]:
            calendario[cod] = r

    psd_commodities = fas("/commodities")
    nomes_attr = {
        a["attributeId"]: a["attributeName"].strip()
        for a in fas("/commodityAttributes")
    }
    esr_commodities = {c["commodityName"]: c["commodityCode"] for c in fas("/commodities", "esr")}

    def semanas(registros: list) -> dict:
        """Agrega os países por semana: {data: {embarques, acumulado, aberto}}."""
        out = {}
        for r in registros:
            dia = (r.get("weekEndingDate") or "")[:10]
            if not dia:
                continue
            a = out.setdefault(dia, {"embarques": 0, "acumulado": 0, "aberto": 0})
            a["embarques"] += r.get("weeklyExports") or 0
            a["acumulado"] += r.get("accumulatedExports") or 0
            a["aberto"] += r.get("outstandingSales") or 0
        return out

    itens, failed = [], []
    ano_esr = None
    for chave, nome_esr, casa_psd in ESR_ALVOS:
        codigo = esr_commodities.get(nome_esr)
        if not codigo:
            failed.append(f"{chave}: '{nome_esr}' não está no catálogo do ESR")
            continue
        try:
            atual, ano = None, None
            for tentativa in (dt.date.today().year + 1, dt.date.today().year):
                dados = fas(
                    f"/exports/commodityCode/{codigo}/allCountries/marketYear/{tentativa}",
                    "esr",
                    timeout=60,
                )
                if dados:
                    atual, ano = dados, tentativa
                    break
            if not atual:
                failed.append(f"{chave}: ESR sem registros")
                continue
            ano_esr = ano_esr or ano

            porsem = semanas(atual)
            datas = sorted(porsem)
            ult = datas[-1]
            cur = porsem[ult]

            # mesma semana do ano anterior, pela data equivalente
            anterior = fas(
                f"/exports/commodityCode/{codigo}/allCountries/marketYear/{ano - 1}",
                "esr",
                timeout=60,
            )
            ant = semanas(anterior)
            alvo_ant = (dt.date.fromisoformat(ult) - dt.timedelta(days=364)).isoformat()
            eq = min(ant, key=lambda d: abs(
                (dt.date.fromisoformat(d) - dt.date.fromisoformat(alvo_ant)).days
            )) if ant else None

            # ritmo: quanto do ano-safra já passou, pelo calendário do próprio ESR
            cal = calendario.get(codigo)
            fracao = None
            if cal:
                ini = dt.date.fromisoformat(cal["marketYearStart"][:10])
                fim = dt.date.fromisoformat(cal["marketYearEnd"][:10])
                # o calendário publicado é o do ano-safra novo; desloca para o
                # ano-safra efetivamente carregado
                desloca = ano - cal["marketYear"]
                ini = ini.replace(year=ini.year + desloca)
                fim = fim.replace(year=fim.year + desloca)
                total = (fim - ini).days
                if total > 0:
                    fracao = min(
                        1.0, max(0.0, (dt.date.fromisoformat(ult) - ini).days / total)
                    )

            # o ano-safra do PSD é o do ESR menos um
            meta, nome_psd = psd_export_us(psd_commodities, casa_psd, nomes_attr, ano - 1)

            ultimas = datas[-4:]
            media4 = sum(porsem[d]["embarques"] for d in ultimas) / len(ultimas)
            compromissos = cur["acumulado"] + cur["aberto"]

            item = {
                "chave": chave,
                "nome": {"soja": "Soja", "farelo": "Farelo de soja",
                         "oleo": "Óleo de soja", "milho": "Milho"}[chave],
                "semana": round(cur["embarques"]),
                "semana_ref": ult,
                "semana_anterior": round(porsem[datas[-2]]["embarques"]) if len(datas) > 1 else None,
                "media_4s": round(media4),
                "semana_ano_passado": round(ant[eq]["embarques"]) if eq else None,
                "acumulado": round(cur["acumulado"]),
                "acumulado_ano_passado": round(ant[eq]["acumulado"]) if eq else None,
                "em_aberto": round(cur["aberto"]),
                "compromissos": round(compromissos),
                "meta": round(meta) if meta else None,
                "meta_fonte": nome_psd,
                "safra": f"{ano - 1}/{str(ano)[-2:]}",
                "fracao_safra": round(fracao, 3) if fracao is not None else None,
                "serie": [
                    {"data": d, "v": round(porsem[d]["embarques"])}
                    for d in datas[-SEMANAS_SERIE:]
                ],
            }
            if meta:
                item["pct_meta"] = round(cur["acumulado"] / meta * 100, 1)
                item["pct_compromissos"] = round(compromissos / meta * 100, 1)
                if fracao:
                    item["indice_ritmo"] = round(
                        (cur["acumulado"] / meta) / fracao * 100, 1
                    )
                semanas_restantes = max(1, round((1 - (fracao or 0)) * 52))
                item["necessario_semanal"] = round(
                    max(0, meta - cur["acumulado"]) / semanas_restantes
                )

            if chave == "soja":
                porpais = {}
                for r in atual:
                    if (r.get("weekEndingDate") or "")[:10] != ult:
                        continue
                    porpais[r["countryCode"]] = r.get("accumulatedExports") or 0
                nomes_pais = {
                    c["countryCode"]: c["countryName"].strip()
                    for c in fas("/countries", "esr")
                }
                top = sorted(porpais.items(), key=lambda kv: -kv[1])[:8]
                item["destinos"] = [
                    {
                        "pais": nomes_pais.get(cod, str(cod)).title(),
                        "acumulado": round(v),
                        "pct": round(v / cur["acumulado"] * 100, 1) if cur["acumulado"] else None,
                        "china": cod == CHINA,
                    }
                    for cod, v in top if v > 0
                ]
            itens.append(item)
        except Exception as e:  # noqa: BLE001
            failed.append(f"{chave}: {e}")

    if not itens:
        raise RuntimeError("; ".join(failed) or "ESR não devolveu nada")
    return {
        "updated_at": now_iso(),
        "unidade": "toneladas",
        "itens": itens,
        "failed": failed,
    }


# ---------------------------------------------------------------- basis

BU_POR_SACA = 60 / 27.2155  # bushels de soja em uma saca de 60 kg
HIST = ROOT / "data" / "basis_history.json"
HIST_MAX = 500


MESES_PT = {
    "jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6,
    "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12,
}


def mes_embarque(rotulo: str) -> tuple:
    """'Agosto/26' -> (2026, 8)."""
    nome, _, ano = rotulo.partition("/")
    return (2000 + int(ano), MESES_PT.get(nome.strip()[:3].lower(), 1))


def mes_contrato(rotulo: str) -> tuple:
    """'set/26' -> (2026, 9)."""
    nome, _, ano = rotulo.partition("/")
    return (2000 + int(ano), MESES_PT.get(nome.strip()[:3].lower(), 1))


def collect_basis(sections: dict) -> dict:
    """Decompõe o preço que o produtor recebe em paridade de exportação + basis.

    O cliente não recebe em ¢/bushel: recebe o indicador em R$/saca, que é
    a paridade (CBOT + prêmio de porto, convertida pelo câmbio) mais um
    basis. O basis é estacionário e sazonal — função de frete, ritmo de
    venda e capacidade portuária — e é a parte modelável do preço.
    """
    fx = sections.get("fx") or {}
    quotes = sections.get("quotes") or {}
    cepea = sections.get("cepea") or {}

    usd = (fx.get("usdbrl") or {}).get("bid")
    zs = next(
        (q.get("price") for q in quotes.get("items", []) if q.get("symbol") == "ZS=F"),
        None,
    )
    if not usd or not zs:
        raise RuntimeError("faltam dólar ou ZS=F para decompor o preço")

    indicadores = cepea.get("indicadores") or []
    porto = next((i for i in indicadores if "Paranaguá" in i["nome"]), None)
    if not porto:
        raise RuntimeError("indicador de Paranaguá indisponível")

    premios = cepea.get("premio_paranagua") or []
    premio = premios[0] if premios else None
    premio_cents = (premio or {}).get("cents_bu") or 0.0
    # Sem prêmio não existe paridade de exportação: o cálculo degenera em
    # "indicador menos Chicago puro", que dá basis de +9% e parece um sinal
    # forte de mercado quando é só um campo faltando. Melhor não publicar e
    # deixar a seção herdar o último valor bom — o Notícias Agrícolas já
    # removeu essa tabela da página no meio de um dia.
    if not premio_cents:
        raise RuntimeError("prêmio de Paranaguá indisponível — basis não calculado")

    # O prêmio é cotado por mês de embarque e precifica contra o contrato
    # CBOT vigente naquele embarque — não contra o primeiro vencimento.
    contrato = None
    if premio:
        alvo = mes_embarque(premio["mes"])
        for c in curva_de(sections, "soja").get("contracts", []):
            if c.get("price") and mes_contrato(c["label"]) >= alvo:
                contrato = c
                break
    if contrato:
        zs = contrato["price"]

    def cents_para_saca(cents: float) -> float:
        return (cents / 100) * BU_POR_SACA * usd

    def saca_para_cents(brl: float) -> float:
        return (brl / usd) / BU_POR_SACA * 100

    flat = cents_para_saca(zs)
    paridade = cents_para_saca(zs + premio_cents)
    basis_brl = round(porto["valor"] - paridade, 2)

    out = {
        "updated_at": now_iso(),
        "cambio": usd,
        "cbot_cents": zs,
        "premio_cents": premio_cents,
        "premio_mes": premio["mes"] if premio else None,
        "contrato": contrato["label"] if contrato else "1º vencimento",
        "flat_brl_saca": round(flat, 2),
        "paridade_brl_saca": round(paridade, 2),
        "indicador_brl_saca": porto["valor"],
        "indicador_data": porto["data"],
        "basis_porto": {
            "brl_saca": basis_brl,
            "cents_bu": round(saca_para_cents(basis_brl), 1),
            "pct": round(basis_brl / paridade * 100, 1) if paridade else None,
        },
    }

    # Basis do interior: desconto da praça contra o porto, que é essencialmente
    # frete até Paranaguá mais o poder de barganha local.
    interior = []
    for f in cepea.get("fisico", []):
        if f["praca"].startswith("Porto"):
            continue
        d = round(f["valor"] - porto["valor"], 2)
        interior.append(
            {
                "praca": f["praca"],
                "valor": f["valor"],
                "basis_brl_saca": d,
                "basis_cents_bu": round(saca_para_cents(d), 1),
            }
        )
    if interior:
        interior.sort(key=lambda x: x["basis_brl_saca"])
        out["basis_interior"] = interior

    out["historico"] = atualiza_historico(out)
    return out


def atualiza_historico(basis: dict) -> list:
    """Acumula uma observação diária do basis — a série que o modelo vai usar.

    Não existe histórico gratuito do indicador para reconstruir o passado,
    então a série é construída daqui para a frente, uma observação por dia
    de referência do indicador.
    """
    try:
        serie = json.loads(HIST.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        serie = []

    registro = {
        "data": basis["indicador_data"],
        "cbot_cents": basis["cbot_cents"],
        "cambio": basis["cambio"],
        "paridade": basis["paridade_brl_saca"],
        "indicador": basis["indicador_brl_saca"],
        "basis": basis["basis_porto"]["brl_saca"],
    }
    serie = [s for s in serie if s.get("data") != registro["data"]]
    serie.append(registro)
    serie.sort(key=lambda s: tuple(reversed(s["data"].split("/"))))
    serie = serie[-HIST_MAX:]

    HIST.parent.mkdir(parents=True, exist_ok=True)
    HIST.write_text(
        json.dumps(serie, ensure_ascii=False, indent=1), encoding="utf-8"
    )
    return serie


ROLLS = ROOT / "data" / "roll_history.json"


def registra_rolagem(quotes: dict) -> None:
    """Anota qual contrato estava por trás de cada contínuo, por dia.

    Retroajustar a série emendada (para que as variações de mais de um dia
    parem de incluir o degrau da troca de contrato) exige saber em que dia
    cada rolagem aconteceu. Ninguém publica isso de graça, então acumulamos
    daqui para a frente — mesmo caminho da série do basis.
    """
    try:
        serie = json.loads(ROLLS.read_text(encoding="utf-8"))
    except Exception:  # noqa: BLE001
        serie = []

    hoje = dt.date.today().isoformat()
    atual = {
        i["symbol"]: i["contrato"]
        for i in (quotes or {}).get("items", [])
        if i.get("contrato")
    }
    if not atual:
        return

    serie = [s for s in serie if s.get("data") != hoje]
    serie.append({"data": hoje, "contratos": atual})
    serie.sort(key=lambda s: s["data"])
    serie = serie[-HIST_MAX:]

    ROLLS.parent.mkdir(parents=True, exist_ok=True)
    ROLLS.write_text(
        json.dumps(serie, ensure_ascii=False, indent=1), encoding="utf-8"
    )


# ---------------------------------------------------------------- main

COLLECTORS = {
    "quotes": collect_quotes,
    "fx": collect_fx,
    "rates": collect_rates,
    "cepea": collect_cepea,
    "safra": collect_safra,
    "curve": collect_curve,
    "options": collect_options,
    "weather": collect_weather,
    "news": collect_news,
    "sd": collect_sd,
    "embarques": collect_embarques,
}

DERIVED = {"basis": collect_basis}


def main() -> int:
    old_sections = {}
    if OUT.exists():
        try:
            old_sections = json.loads(OUT.read_text(encoding="utf-8")).get(
                "sections", {}
            )
        except Exception:  # noqa: BLE001
            pass

    sections, errors = {}, []
    for name, collector in COLLECTORS.items():
        try:
            sections[name] = collector()
            print(f"[ok]   {name}")
        except Exception as e:  # noqa: BLE001
            errors.append({"section": name, "error": str(e)})
            sections[name] = old_sections.get(name)
            print(f"[erro] {name}: {e}", file=sys.stderr)

    # Derivados: dependem das seções acima, então rodam depois delas.
    for name, derive in DERIVED.items():
        try:
            sections[name] = derive(sections)
            print(f"[ok]   {name}")
        except Exception as e:  # noqa: BLE001
            errors.append({"section": name, "error": str(e)})
            sections[name] = old_sections.get(name)
            print(f"[erro] {name}: {e}", file=sys.stderr)

    anota_cadencia(sections)

    try:
        registra_rolagem(sections.get("quotes") or {})
    except Exception as e:  # noqa: BLE001 - histórico auxiliar, não trava a coleta
        print(f"[erro] roll_history: {e}", file=sys.stderr)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        json.dumps(
            {"generated_at": now_iso(), "sections": sections, "errors": errors},
            ensure_ascii=False,
            separators=(",", ":"),
        ),
        encoding="utf-8",
    )
    print(f"gravado {OUT} ({len(errors)} erro(s))")
    # Só falha o job se absolutamente nada funcionou
    return 1 if len(errors) == len(COLLECTORS) + len(DERIVED) else 0


if __name__ == "__main__":
    sys.exit(main())
