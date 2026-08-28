"""Sonda temporária: fontes de custo para o item B4 do plano.

Duas pressões que o painel ainda não mede:

  fertilizante — entra na relação de troca (quantas sacas de soja compram uma
                 tonelada de ureia), que é como o produtor decide compra de
                 insumo. Candidata: Pink Sheet do Banco Mundial, mensal e
                 aberto.
  diesel       — explica o basis de interior via frete. Candidata: série
                 histórica da ANP. O Brent já coletado é proxy imediato, mas
                 não captura política de preço da Petrobras nem câmbio.

Objetivo: descobrir quais URLs respondem, em que formato, e com que colunas —
para decidir o que dá para implementar sem inventar dado.

Rodar pelo workflow `probe.yml` e ler os logs. Remover depois.
"""

import io
import re
import time

import requests

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "pt-BR,pt;q=0.9,en;q=0.8",
}

PINK_SHEET = [
    ("xlsx mensal",
     "https://thedocs.worldbank.org/en/doc/5d903e848db1d1b83e0ec8f744e55570-0350012021/"
     "related/CMO-Historical-Data-Monthly.xlsx"),
    ("página CMO",
     "https://www.worldbank.org/en/research/commodity-markets"),
]

ANP = [
    ("dados abertos ANP",
     "https://www.gov.br/anp/pt-br/centrais-de-conteudo/dados-abertos/"
     "serie-historica-de-precos-de-combustiveis"),
    ("CKAN dados.gov.br",
     "https://dados.gov.br/api/3/action/package_search?q=anp+combustiveis&rows=5"),
]


def tenta(rotulo: str, url: str):
    try:
        r = requests.get(url, headers=HEADERS, timeout=60, allow_redirects=True)
        ct = r.headers.get("Content-Type", "")[:60]
        print(f"  {r.status_code}  {len(r.content):>9,} B  {ct:<40} {rotulo}")
        return r if r.ok else None
    except Exception as e:  # noqa: BLE001
        print(f"  ERRO  {rotulo}: {str(e)[:110]}")
        return None


def inspeciona_xlsx(r) -> None:
    """Abas e primeiras linhas — para saber onde estão ureia, DAP e KCl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("     (openpyxl não instalado)")
        return
    try:
        wb = load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
        print(f"     abas: {wb.sheetnames}")
        for aba in wb.sheetnames:
            ws = wb[aba]
            print(f"     --- {aba} ({ws.max_row}x{ws.max_column}) ---")
            for i, linha in enumerate(ws.iter_rows(max_row=8, values_only=True)):
                cels = [str(c)[:22] for c in linha[:12] if c is not None]
                if cels:
                    print(f"       {i}: {' | '.join(cels)}")
            # onde estão os fertilizantes?
            for linha in ws.iter_rows(max_row=8, values_only=True):
                achados = [
                    str(c) for c in linha
                    if c and re.search(r"urea|dap|potassium|phosphate|fertil",
                                       str(c), re.I)
                ]
                if achados:
                    print(f"       fertilizantes nesta linha: {achados[:8]}")
            if aba != wb.sheetnames[-1]:
                print()
    except Exception as e:  # noqa: BLE001
        print(f"     não abriu como xlsx: {str(e)[:110]}")


def main() -> None:
    print(f"{'=' * 78}\nFERTILIZANTE — Pink Sheet do Banco Mundial\n{'=' * 78}")
    for rotulo, url in PINK_SHEET:
        r = tenta(rotulo, url)
        if r and "sheet" in r.headers.get("Content-Type", "").lower():
            inspeciona_xlsx(r)
        elif r and rotulo.startswith("página"):
            # procura o link do xlsx na página, caso a URL direta tenha mudado
            links = re.findall(r'href="([^"]*\.xlsx?)"', r.text, re.I)
            print(f"     links .xls(x) na página: {links[:6] or 'nenhum'}")
        time.sleep(2)

    print(f"\n{'=' * 78}\nDIESEL — ANP\n{'=' * 78}")
    for rotulo, url in ANP:
        r = tenta(rotulo, url)
        if not r:
            continue
        if "json" in r.headers.get("Content-Type", ""):
            try:
                res = r.json()["result"]["results"]
                for p in res[:5]:
                    print(f"     · {p.get('title','')[:70]}")
                    for rec in (p.get("resources") or [])[:3]:
                        print(f"         {rec.get('format','?'):6} {rec.get('url','')[:90]}")
            except Exception as e:  # noqa: BLE001
                print(f"     json inesperado: {str(e)[:90]}")
        else:
            links = re.findall(r'href="([^"]*\.(?:csv|xlsx?|zip))"', r.text, re.I)
            print(f"     arquivos linkados: {links[:8] or 'nenhum'}")
        time.sleep(2)

    print(f"\n{'=' * 78}\nVEREDITO\n{'=' * 78}")
    print("  Implementar só o que respondeu com formato legível.")
    print("  Fonte que exige raspagem de página gov.br muda de endereço e quebra;")
    print("  melhor deixar de fora do que colocar e ver falhar em silêncio.")


if __name__ == "__main__":
    main()
